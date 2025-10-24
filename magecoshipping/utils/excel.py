import sqlite3
import pandas as pd
from pathlib import Path
from datetime import datetime
import re
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from magecoshipping.db.schema import DB_PATH

# Intestazioni fisse e ordine del report (10 colonne)
FULL_HEADERS = [
    "Data Viaggio",
    "Rotta",
    "Tratta",
    "Tipologia di Veicolo (vedi foglio Nomenclatura veicoli imbarcati)",
    "Identificativo cassa mobile",
    "VIN",
    "Targa motrice",
    "Targa rimorchio",
    "Num Mezzi",
    "Costo imponibile quietanzato (pagato)",
]


def export_filtered_excel(filters=None) -> Path:
    """
    Esporta i documenti filtrati (o selezionati) in un file Excel formattato.
    Regole:
      - Solo righe con include = 1
      - Duplicazione righe per targhe multiple
      - Quantità e Num Mezzi sempre = 1
      - Costo unitario = costo_totale / num_targhe
      - Colonne SEMPRE presenti nell'ordine FULL_HEADERS (anche vuote)
    """
    exports_dir = Path(__file__).resolve().parents[1] / "exports"
    exports_dir.mkdir(exist_ok=True)

    output_path = exports_dir / f"Report_Mageco_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # 🔹 Recupera i documenti da esportare
    if filters and "ids" in filters:
        placeholders = ",".join("?" for _ in filters["ids"])
        cur.execute(f"SELECT * FROM documents WHERE id IN ({placeholders}) ORDER BY created_at DESC", filters["ids"])
    else:
        cur.execute("SELECT * FROM documents ORDER BY created_at DESC")

    documents = [dict(row) for row in cur.fetchall()]
    wb_writer = pd.ExcelWriter(output_path, engine="openpyxl")

    for doc in documents:
        doc_id = doc["id"]

        # 🔹 Intestazione documento
        header_data = {
            "Soggetto Certificatore": doc.get("fornitore", ""),
            "Cliente": doc.get("cliente", ""),
            "Partita IVA Cliente": doc.get("piva_cliente", ""),
        }

        # 🔹 Righe documento (solo incluse)
        cur.execute("""
            SELECT descrizione_rigo, tratta, targhe, tipo_veicolo,
                   quantita_fattura, quantita_reale, costo, include
            FROM document_lines
            WHERE document_id = ? AND include = 1
            ORDER BY id
        """, (doc_id,))
        lines = [dict(row) for row in cur.fetchall()]
        if not lines:
            continue

        processed_rows = []
        for r in lines:
            targhe_raw = str(r.get("targhe") or "").strip()
            # split su trattini lunghi, virgole, slash e whitespace
            targhe = [t.strip().upper() for t in re.split(r"[–,/\s]+", targhe_raw) if t.strip()]
            num_targhe = len(targhe) if targhe else 1

            costo_totale = float(r.get("costo") or 0)
            costo_unitario = round(costo_totale / num_targhe, 2)

            if not targhe:
                # nessuna targa -> una riga generica
                r_copy = r.copy()
                r_copy["targhe"] = ""
                r_copy["costo"] = costo_unitario
                r_copy["quantita_reale"] = 1
                r_copy["num_mezzi"] = 1
                processed_rows.append(r_copy)
            else:
                for targa in targhe:
                    r_copy = r.copy()
                    r_copy["targhe"] = targa
                    r_copy["costo"] = costo_unitario
                    r_copy["quantita_reale"] = 1
                    r_copy["num_mezzi"] = 1   # <--- sempre 1 per ogni riga
                    processed_rows.append(r_copy)

        # 🔹 Crea DataFrame formattato
        df = pd.DataFrame(processed_rows)
        if df.empty:
            continue

        # Valori di base
        data_viaggio = doc.get("data_doc", "")  # se in futuro avrai la data per rigo, spostala nel loop sopra
        df["Data Viaggio"] = data_viaggio

        # Rinomina verso le intestazioni finali
        df.rename(columns={
            "tratta": "Tratta",
            "tipo_veicolo": "Tipologia di Veicolo (vedi foglio Nomenclatura veicoli imbarcati)",
            "targhe": "Targa motrice",
            "num_mezzi": "Num Mezzi",
            "costo": "Costo imponibile quietanzato (pagato)"
        }, inplace=True)

        # Colonne non ancora presenti nel DB (al momento) → vuote
        if "Rotta" not in df.columns:
            df["Rotta"] = ""
        if "Identificativo cassa mobile" not in df.columns:
            df["Identificativo cassa mobile"] = ""
        if "VIN" not in df.columns:
            df["VIN"] = ""
        if "Targa rimorchio" not in df.columns:
            df["Targa rimorchio"] = ""

        # Regole: Num Mezzi sempre 1, costo numerico (0.0 se mancante)
        df["Num Mezzi"] = 1
        df["Costo imponibile quietanzato (pagato)"] = pd.to_numeric(
            df.get("Costo imponibile quietanzato (pagato)"), errors="coerce"
        ).fillna(0.0)

        # Allinea e forza l'ordine colonne: sempre le 10 intestazioni
        df = df.reindex(columns=FULL_HEADERS, fill_value="")

        # 🔹 Scrivi nel foglio Excel
        sheet_name = f"{doc.get('cliente', '')[:25]}_{doc_id}"
        df.to_excel(wb_writer, index=False, sheet_name=sheet_name, startrow=6)

        # 🔹 Intestazione cliente / fornitore
        ws = wb_writer.sheets[sheet_name]
        ws["E1"] = header_data["Soggetto Certificatore"]
        ws["E2"] = header_data["Cliente"]
        ws["E3"] = header_data["Partita IVA Cliente"]

        ws["D1"].value = "Soggetto Certificatore"
        ws["D2"].value = "Cliente"
        ws["D3"].value = "Partita Iva Cliente"

        bold = Font(bold=True, color="1F497D")
        for r in range(1, 4):
            ws[f"D{r}"].font = bold

        # 🔹 Stile intestazione tabella (10 colonne)
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center = Alignment(horizontal="center", vertical="center")

        header_row = 7
        for col in range(1, len(FULL_HEADERS) + 1):
            cell = ws.cell(row=header_row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        # 🔹 Larghezze colonne
        widths = [15, 12, 12, 42, 25, 20, 20, 20, 10, 25]  # 10 colonne
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = w

        # 🔹 Bordo sottile su tutta la tabella
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row, min_col=1, max_col=len(FULL_HEADERS)):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        # 🔹 Formati colonne: data / intero / valuta
        data_start = header_row + 1
        data_end = ws.max_row

        # Data Viaggio (col A)
        for r_idx in range(data_start, data_end + 1):
            ws[f"A{r_idx}"].number_format = "DD/MM/YYYY"

        # Num Mezzi (col I)
        for r_idx in range(data_start, data_end + 1):
            ws[f"I{r_idx}"].number_format = "0"

        # Costo (col J)
        for r_idx in range(data_start, data_end + 1):
            ws[f"J{r_idx}"].number_format = "€ #,##0.00"

        # Blocca intestazione
        ws.freeze_panes = "A8"

    wb_writer.close()
    conn.close()
    return output_path
